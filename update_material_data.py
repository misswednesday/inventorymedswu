import os
import shutil
import json
import openpyxl

def update_data():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    print(f"Scanning directory: {base_dir}")
    
    excel_candidates = [f for f in os.listdir(base_dir) if (f.endswith('.xlsx') or f.endswith('.xls')) and not f.startswith('~$')]
    
    selected_excel = None
    for f in excel_candidates:
        if 'วัสดุ' in f or '1กย69' in f or '169' in f:
            selected_excel = f
            break
            
    if not selected_excel and excel_candidates:
        for f in excel_candidates:
            try:
                p = os.path.join(base_dir, f)
                wb_check = openpyxl.load_workbook(p, read_only=True)
                if 'Data' in wb_check.sheetnames:
                    selected_excel = f
                    break
            except:
                pass
                
    if not selected_excel and excel_candidates:
        selected_excel = excel_candidates[0]
        
    if not selected_excel:
        print("Error: No suitable Excel file found!")
        return False
        
    excel_path = os.path.join(base_dir, selected_excel)
    print(f"Reading Excel file: {selected_excel}")
    
    # Image folder
    img_folder_name = 'รูปวัสดุคงคลัง'
    img_folder_path = os.path.join(base_dir, img_folder_name)
    available_images = {}
    if os.path.exists(img_folder_path):
        for f in os.listdir(img_folder_path):
            base, ext = os.path.splitext(f)
            available_images[base] = f
        print(f"Found {len(available_images)} images in {img_folder_name}")
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # Data sheet
    data_sheet = wb['Data'] if 'Data' in wb.sheetnames else wb.worksheets[0]
    items = []
    for r in range(2, data_sheet.max_row + 1):
        mat_code = data_sheet.cell(r, 1).value
        desc = data_sheet.cell(r, 2).value
        
        # Skip if material code or description is empty / None / whitespace
        if mat_code is None or desc is None:
            continue
            
        mat_str = str(mat_code).strip()
        desc_str = str(desc).strip()
        
        if not mat_str or not desc_str:
            continue
            
        bun = data_sheet.cell(r, 3).value
        bun_str = str(bun).strip() if bun is not None else ""
        unrestricted = data_sheet.cell(r, 4).value
        try:
            unrestricted_val = float(unrestricted) if unrestricted is not None else 0.0
            if unrestricted_val.is_integer():
                unrestricted_val = int(unrestricted_val)
        except:
            unrestricted_val = 0
            
        img_file = available_images.get(mat_str, None)
        items.append({
            'id': mat_str,
            'material': mat_str,
            'name': desc_str,
            'unit': bun_str,
            'stock': unrestricted_val,
            'image': f"{img_folder_name}/{img_file}" if img_file else None,
            'has_image': img_file is not None
        })
        
    # Department sheet
    dept_sheet = None
    for s in wb.sheetnames:
        if s != data_sheet.title:
            dept_sheet = wb[s]
            break
            
    departments = []
    if dept_sheet:
        for r in range(1, dept_sheet.max_row + 1):
            val = dept_sheet.cell(r, 1).value
            if val:
                v_str = str(val).strip()
                if v_str and v_str not in departments:
                    departments.append(v_str)
    departments = sorted(departments)
    
    js_content = f"""// Auto-generated data file for MED SWU Inventory & Requisition System
// Generated from: {selected_excel}

const INVENTORY_DATA = {{
  metadata: {{
    generatedAt: "{openpyxl.__name__}",
    totalItems: {len(items)},
    itemsWithImages: {sum(1 for x in items if x['has_image'])},
    totalDepartments: {len(departments)}
  }},
  departments: {json.dumps(departments, ensure_ascii=False, indent=2)},
  items: {json.dumps(items, ensure_ascii=False, indent=2)}
}};

if (typeof module !== 'undefined' && module.exports) {{
  module.exports = INVENTORY_DATA;
}}
"""
    out_js = os.path.join(base_dir, 'materials_data.js')
    with open(out_js, 'w', encoding='utf-8') as f:
        f.write(js_content)
    print(f"Successfully updated {out_js} ({len(items)} items, {len(departments)} departments)")
    return True

if __name__ == '__main__':
    update_data()
