import os
import sys
import json
import time
import openpyxl
from datetime import datetime
from http.server import HTTPServer, SimpleHTTPRequestHandler
import threading

PORT = 8765
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PEER_DIRS = ['i:/ฐานข้อมูลครุภัณฑ์2570', 'I:/วัสดุคงคลัง 2570']

def get_target_excel_paths():
    paths = []
    for d in [BASE_DIR] + PEER_DIRS:
        if os.path.exists(d):
            try:
                excel_files = [f for f in os.listdir(d) if (f.endswith('.xlsx') or f.endswith('.xls')) and not f.startswith('~$')]
                for f in excel_files:
                    if 'วัสดุ' in f or '1กย69' in f or '169' in f:
                        p = os.path.join(d, f)
                        if p not in paths:
                            paths.append(p)
            except:
                pass
    return paths

def write_requisitions_to_excel(reqs_list):
    excel_paths = get_target_excel_paths()
    headers = [
        'วันเวลา',
        'ชื่อภาควิชา/หน่วยงาน',
        'รายการที่เบิก',
        'จำนวน',
        'หน่วยนับ',
        'ชื่อผู้เบิก',
        'สถานะ',
        'รหัสวัสดุ',
        'เลขที่ใบเบิก',
        'ผู้อนุญาต',
        'วัตถุประสงค์'
    ]
    
    rows_data = []
    for r in reqs_list:
        status_text = 'รอการอนุมัติ'
        if r.get('status') == 'approved':
            status_text = 'อนุมัติแล้ว'
        elif r.get('status') == 'rejected':
            status_text = f"ไม่อนุมัติ ({r.get('rejectionReason') or 'ไม่อนุมัติ'})"
            
        for it in r.get('items', []):
            rows_data.append([
                r.get('formattedDate', ''),
                r.get('department', ''),
                it.get('name', ''),
                it.get('qty', 0),
                it.get('unit', ''),
                r.get('requesterName', ''),
                status_text,
                it.get('material', ''),
                r.get('id', ''),
                r.get('approvedBy') or ('หัวหน้างานพัสดุ' if r.get('status') == 'approved' else '-'),
                r.get('remark', '-')
            ])

    success_count = 0
    for p in excel_paths:
        try:
            wb = openpyxl.load_workbook(p)
            sheet_name = 'การเบิก'
            
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
                
            ws = wb.create_sheet(title=sheet_name)
            ws.append(headers)
            
            for row in rows_data:
                ws.append(row)
                
            col_widths = {
                'A': 22, 'B': 30, 'C': 35, 'D': 10, 'E': 10,
                'F': 25, 'G': 18, 'H': 15, 'I': 22, 'J': 25, 'K': 30
            }
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width
                
            wb.save(p)
            print(f"[{datetime.now().strftime('%H:%M:%S')}] บันทึกสำเร็จ: เขียนข้อมูล {len(rows_data)} แถวลงใน Sheet 'การเบิก' ของ {os.path.basename(p)}")
            success_count += 1
        except Exception as e:
            print(f"[{datetime.now().strftime('%H:%M:%S')}] คำเตือน: ไม่สามารถเขียนลง {p} ({e})")

    # Save JSON backup
    backup_file = os.path.join(BASE_DIR, 'requisitions_data.json')
    try:
        with open(backup_file, 'w', encoding='utf-8') as f:
            json.dump(reqs_list, f, ensure_ascii=False, indent=2)
    except:
        pass

    return True, len(rows_data)

class UnifiedHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=BASE_DIR, **kwargs)

    def _set_cors_headers(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.send_header('Access-Control-Allow-Private-Network', 'true')

    def do_OPTIONS(self):
        self.send_response(200)
        self._set_cors_headers()
        self.end_headers()

    def do_GET(self):
        if self.path.startswith('/api/get_requisitions'):
            backup_file = os.path.join(BASE_DIR, 'requisitions_data.json')
            data = []
            if os.path.exists(backup_file):
                try:
                    with open(backup_file, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                except:
                    pass
            self.send_response(200)
            self._set_cors_headers()
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.end_headers()
            self.wfile.write(json.dumps({'requisitions': data}, ensure_ascii=False).encode('utf-8'))
        else:
            super().do_GET()

    def do_POST(self):
        if self.path.startswith('/api/save_requisitions'):
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length)
            try:
                payload = json.loads(body.decode('utf-8'))
                reqs = payload.get('requisitions', [])
                ok, row_count = write_requisitions_to_excel(reqs)
                
                resp = {
                    'success': ok,
                    'message': f'บันทึกข้อมูล {row_count} รายการลงใน Sheet การเบิก ของไฟล์ Excel เรียบร้อยแล้ว',
                    'rowsCount': row_count
                }
                self.send_response(200)
                self._set_cors_headers()
                self.send_header('Content-Type', 'application/json; charset=utf-8')
                self.end_headers()
                self.wfile.write(json.dumps(resp, ensure_ascii=False).encode('utf-8'))
            except Exception as e:
                self.send_response(500)
                self._set_cors_headers()
                self.send_header('Content-Type', 'application/json; charset=utf-8')
                self.end_headers()
                self.wfile.write(json.dumps({'success': False, 'error': str(e)}, ensure_ascii=False).encode('utf-8'))
        else:
            self.send_response(404)
            self.end_headers()

    def end_headers(self):
        self._set_cors_headers()
        super().end_headers()

    def log_message(self, format, *args):
        pass

def parse_and_update(excel_path, base_dir):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] ตรวจพบการแก้ไขไฟล์: {os.path.basename(excel_path)}")
    
    img_folder_name = 'รูปวัสดุคงคลัง'
    img_folder_path = os.path.join(base_dir, img_folder_name)
    available_images = {}
    if os.path.exists(img_folder_path):
        try:
            for f in os.listdir(img_folder_path):
                base, ext = os.path.splitext(f)
                available_images[base] = f
        except:
            pass

    wb = None
    for attempt in range(5):
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            break
        except:
            time.sleep(0.5)
            
    if not wb:
        return False
        
    data_sheet = wb['Data'] if 'Data' in wb.sheetnames else wb.worksheets[0]
    items = []
    for r in range(2, data_sheet.max_row + 1):
        mat_code = data_sheet.cell(r, 1).value
        desc = data_sheet.cell(r, 2).value
        if mat_code is None or desc is None:
            continue
        mat_str = str(mat_code).strip()
        desc_str = str(desc).strip()
        if not mat_str or not desc_str:
            continue
        
        bun = data_sheet.cell(r, 3).value
        bun_str = str(bun).strip() if bun is not None else ""
        unrestricted = data_sheet.cell(r, 4).value
        try:
            unrestricted_val = float(unrestricted) if unrestricted is not None else 0.0
            if unrestricted_val.is_integer():
                unrestricted_val = int(unrestricted_val)
        except:
            unrestricted_val = 0
            
        img_file = available_images.get(mat_str, None)
        items.append({
            'id': mat_str,
            'material': mat_str,
            'name': desc_str,
            'unit': bun_str,
            'stock': unrestricted_val,
            'image': f"{img_folder_name}/{img_file}" if img_file else None,
            'has_image': img_file is not None
        })
        
    dept_sheet = None
    for s in wb.sheetnames:
        if s != data_sheet.title and s != 'การเบิก':
            dept_sheet = wb[s]
            break
            
    departments = []
    if dept_sheet:
        for r in range(1, dept_sheet.max_row + 1):
            val = dept_sheet.cell(r, 1).value
            if val:
                v_str = str(val).strip()
                if v_str and v_str not in departments:
                    departments.append(v_str)
    departments = sorted(departments)
    
    timestamp = int(time.time() * 1000)
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    js_content = f"""// Auto-generated data file for MED SWU Inventory & Requisition System
// Generated from: {os.path.basename(excel_path)} at {now_str}

const INVENTORY_DATA = {{
  metadata: {{
    generatedAt: "{now_str}",
    timestamp: {timestamp},
    totalItems: {len(items)},
    itemsWithImages: {sum(1 for x in items if x['has_image'])},
    totalDepartments: {len(departments)}
  }},
  departments: {json.dumps(departments, ensure_ascii=False, indent=2)},
  items: {json.dumps(items, ensure_ascii=False, indent=2)}
}};

if (typeof module !== 'undefined' && module.exports) {{
  module.exports = INVENTORY_DATA;
}}
"""
    out_js = os.path.join(base_dir, 'materials_data.js')
    with open(out_js, 'w', encoding='utf-8') as f:
        f.write(js_content)
        
    for p_dir in PEER_DIRS:
        if os.path.exists(p_dir) and p_dir != base_dir:
            try:
                with open(os.path.join(p_dir, 'materials_data.js'), 'w', encoding='utf-8') as f:
                    f.write(js_content)
            except:
                pass
                
    print(f"[{datetime.now().strftime('%H:%M:%S')}] ซิงก์ข้อมูลสต๊อกสดแล้ว ({len(items)} รายการ)")
    return True

def start_http_server():
    try:
        server = HTTPServer(('0.0.0.0', PORT), UnifiedHandler)
        print(f"[{datetime.now().strftime('%H:%M:%S')}] ระบบเซิร์ฟเวอร์เปิดทำงานที่ http://localhost:{PORT}")
        server.serve_forever()
    except Exception as e:
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Server Notice: {e}")

def watch_loop():
    print("=" * 65)
    print("  MED SWU - ระบบจัดการสต๊อกและบันทึกใบเบิกลง Excel อัตโนมัติ")
    print(f"  โฟลเดอร์ทำงาน: {BASE_DIR}")
    print(f"  เปิดใช้งานเว็บที่: http://localhost:{PORT}")
    print("=" * 65)
    
    server_thread = threading.Thread(target=start_http_server, daemon=True)
    server_thread.start()
    
    last_mtime = 0
    while True:
        try:
            excel_files = [f for f in os.listdir(BASE_DIR) if (f.endswith('.xlsx') or f.endswith('.xls')) and not f.startswith('~$')]
            curr_excel = None
            for f in excel_files:
                if 'วัสดุ' in f or '1กย69' in f or '169' in f:
                    curr_excel = os.path.join(BASE_DIR, f)
                    break
            if not curr_excel and excel_files:
                curr_excel = os.path.join(BASE_DIR, excel_files[0])
                
            if curr_excel and os.path.exists(curr_excel):
                mtime = os.path.getmtime(curr_excel)
                if mtime != last_mtime:
                    parse_and_update(curr_excel, BASE_DIR)
                    last_mtime = mtime
            time.sleep(1.5)
        except KeyboardInterrupt:
            print("\nหยุดการทำงาน")
            break
        except Exception:
            time.sleep(2)

if __name__ == '__main__':
    watch_loop()
